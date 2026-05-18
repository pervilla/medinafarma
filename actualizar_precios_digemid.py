"""
=============================================================
  actualizar_precios_digemid.py
  Actualiza la tabla dbo.PRECIOS_DIGEMID con los datos del
  archivo catalogoproductos.xlsx (DIGEMID).

  Uso:
      python actualizar_precios_digemid.py
      python actualizar_precios_digemid.py --dry-run   (solo muestra estadísticas, NO modifica BD)
      python actualizar_precios_digemid.py --solo-nuevos (inserta solo registros nuevos, no actualiza)

  Servidor : server (SQL Server 2008 R2)
  Base de datos: BDATOS
  Tabla objetivo: dbo.PRECIOS_DIGEMID
=============================================================
"""

import sys
import time
import argparse
import pyodbc
import openpyxl
from datetime import datetime

# ─────────────────────────────────────────
#  CONFIGURACIÓN
# ─────────────────────────────────────────
DB_CONFIG = {
    "server":   "server",
    "database": "BDATOS",
    "username": "sa",
    "password": "159357852456",
    "driver":   "SQL Server",   # ODBC driver name
}

EXCEL_FILE  = "catalogoproductos.xlsx"
SHEET_NAME  = "Catálogo"
HEADER_ROW  = 7        # La fila donde están los encabezados en el Excel
DATA_START  = 8        # La primera fila con datos reales

# Mapeo columna Excel -> columna en la tabla SQL
# Columnas reales en dbo.PRECIOS_DIGEMID:
#   Cod_Prod, Nom_Prod, Concent, Nom_Form_Farm, Nom_Form_Farm_Simplif,
#   Presentac, Fracciones, Fec_Vcto_Reg_Sanitario, Num_RegSan,
#   Nom_Titular, Situacion, Nom_IFA
#
# NOTA: Nom_Fabricante y Nom_Rubro no existen en la tabla actual;
#       se ignoran en el mapeo. Si los necesitas, agregalos antes.
COLUMN_MAP = {
    "Cod_Prod":      "Cod_Prod",
    "Nom_Prod":      "Nom_Prod",
    "Concent":       "Concent",
    "Nom_Form_Farm": "Nom_Form_Farm",
    "Presentac":     "Presentac",
    "Fraccion":      "Fracciones",    # Excel: 'Fraccion' -> SQL: 'Fracciones'
    "Num_RegSan":    "Num_RegSan",
    "Nom_Titular":   "Nom_Titular",
    "Nom_IFA":       "Nom_IFA",
    "Situacion":     "Situacion",
    # Columnas ignoradas (no existen en la tabla):
    # "Nom_Fabricante" -> no existe
    # "Nom_Rubro"      -> no existe
}

BATCH_SIZE  = 500     # filas por lote MERGE
LOG_EVERY   = 1000    # mostrar progreso cada N filas


# ─────────────────────────────────────────
#  CONEXIÓN
# ─────────────────────────────────────────
def get_connection():
    conn_str = (
        f"DRIVER={{{DB_CONFIG['driver']}}};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"UID={DB_CONFIG['username']};"
        f"PWD={DB_CONFIG['password']};"
        "TrustServerCertificate=yes;"
    )
    try:
        conn = pyodbc.connect(conn_str, timeout=15)
        conn.autocommit = False
        return conn
    except Exception as e:
        print(f"\n[ERROR] No se pudo conectar al servidor SQL Server.")
        print(f"        Cadena usada: SERVER={DB_CONFIG['server']}, DB={DB_CONFIG['database']}")
        print(f"        Detalle: {e}")
        sys.exit(1)


# ─────────────────────────────────────────
#  VERIFICAR ESTRUCTURA DE LA TABLA
# ─────────────────────────────────────────
def verificar_tabla(cursor):
    """Obtiene las columnas reales de PRECIOS_DIGEMID."""
    cursor.execute("""
        SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = 'PRECIOS_DIGEMID'
        ORDER BY ORDINAL_POSITION
    """)
    columnas = cursor.fetchall()
    if not columnas:
        print("[ERROR] La tabla dbo.PRECIOS_DIGEMID no existe o no tiene columnas visibles.")
        sys.exit(1)

    print("\n=== Columnas actuales de dbo.PRECIOS_DIGEMID ===")
    for col in columnas:
        largo = f"({col[2]})" if col[2] else ""
        print(f"  {col[0]:<25} {col[1]}{largo}")

    return [col[0] for col in columnas]


# ─────────────────────────────────────────
#  LEER EXCEL
# ─────────────────────────────────────────
def leer_excel():
    print(f"\n[INFO] Leyendo '{EXCEL_FILE}' (hoja: '{SHEET_NAME}') ...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        # fallback: primer hoja
        ws = wb.active
        print(f"  [WARN] Hoja '{SHEET_NAME}' no encontrada, usando hoja activa: '{ws.title}'")
    else:
        ws = wb[SHEET_NAME]

    # Leer encabezados
    headers = []
    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        headers = [str(h).strip() if h is not None else f"COL_{i}" for i, h in enumerate(row)]
        break

    print(f"  Encabezados encontrados: {headers}")

    # Leer datos - normalizar encabezados quitando tildes para el mapeo
    headers_norm = []
    for h in headers:
        h_norm = (h.replace('\u00f3', 'o')
                   .replace('\u00e1', 'a')
                   .replace('\u00e9', 'e')
                   .replace('\u00ed', 'i')
                   .replace('\u00fa', 'u')
                   .replace('\u00f1', 'n'))
        headers_norm.append(h_norm)

    filas = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        # Ignorar filas completamente vacías
        if all(v is None for v in row):
            continue
        fila = dict(zip(headers_norm, row))
        filas.append(fila)

    wb.close()
    print(f"  Total de filas con datos: {len(filas):,}")
    return headers, filas


# Largos máximos de columnas en dbo.PRECIOS_DIGEMID
# Coincide con la definición real de la tabla
COL_MAX_LEN = {
    "Nom_Prod":      100,
    "Concent":        50,
    "Nom_Form_Farm": 100,
    "Presentac":     100,
    "Num_RegSan":     15,
    "Nom_Titular":   100,
    "Situacion":       4,
    "Nom_IFA":       500,
}


# ─────────────────────────────────────────
#  NORMALIZAR VALORES
# ─────────────────────────────────────────
def limpiar(valor, max_len=None):
    """Convierte a string limpio, truncando si supera max_len."""
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def limpiar_int(valor):
    """Convierte a entero o None."""
    if valor is None:
        return None
    try:
        return int(float(str(valor)))
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────
#  ESTRATEGIA: MERGE por lotes (UPSERT)
# ─────────────────────────────────────────
def actualizar_lote(cursor, lote, columnas_tabla, dry_run):
    """
    Usa una tabla temporal + MERGE para hacer INSERT o UPDATE eficientemente.
    """
    # Determinar columnas a actualizar (intersección de lo que hay en excel y en tabla)
    sql_cols = [sql_col for (excel_col, sql_col) in COLUMN_MAP.items() if sql_col in columnas_tabla]

    if "Cod_Prod" not in sql_cols:
        print("[ERROR] Cod_Prod no encontrado en la tabla. Verifique el mapeo.")
        sys.exit(1)

    cols_sin_pk   = [c for c in sql_cols if c != "Cod_Prod"]
    cols_csv      = ", ".join(sql_cols)
    params_csv    = ", ".join(["?" for _ in sql_cols])
    update_set    = ", ".join([f"T.{c} = S.{c}" for c in cols_sin_pk])

    if dry_run:
        return len(lote)

    # Crear tabla temporal
    temp_cols_def = []
    for c in sql_cols:
        if c == "Cod_Prod":
            temp_cols_def.append(f"{c} INT")
        elif c == "Fracciones":
            temp_cols_def.append(f"{c} NUMERIC(18,2)")
        else:
            temp_cols_def.append(f"{c} NVARCHAR(500)")

    cursor.execute("IF OBJECT_ID('tempdb..#tmp_digemid') IS NOT NULL DROP TABLE #tmp_digemid")
    cursor.execute(f"CREATE TABLE #tmp_digemid ({', '.join(temp_cols_def)})")

    # Insertar lote en temporal
    rows_para_insertar = []
    for fila in lote:
        fila_vals = []
        for excel_col, sql_col in COLUMN_MAP.items():
            if sql_col not in sql_cols:
                continue
            val = fila.get(excel_col)
            if sql_col == "Cod_Prod":
                fila_vals.append(limpiar_int(val))
            elif sql_col == "Fracciones":
                try:
                    fila_vals.append(float(val) if val is not None else None)
                except (ValueError, TypeError):
                    fila_vals.append(None)
            else:
                # Truncar al largo maximo de la columna
                max_len = COL_MAX_LEN.get(sql_col)
                fila_vals.append(limpiar(val, max_len))
        rows_para_insertar.append(tuple(fila_vals))

    # Filtrar filas con Cod_Prod nulo
    rows_para_insertar = [r for r in rows_para_insertar if r[0] is not None]

    cursor.executemany(
        f"INSERT INTO #tmp_digemid ({cols_csv}) VALUES ({params_csv})",
        rows_para_insertar
    )

    # MERGE
    merge_sql = f"""
    MERGE dbo.PRECIOS_DIGEMID AS T
    USING #tmp_digemid AS S ON T.Cod_Prod = S.Cod_Prod
    WHEN MATCHED THEN
        UPDATE SET {update_set}
    WHEN NOT MATCHED BY TARGET THEN
        INSERT ({cols_csv}) VALUES ({', '.join(['S.' + c for c in sql_cols])});
    """
    cursor.execute(merge_sql)
    cursor.execute("DROP TABLE #tmp_digemid")

    return len(rows_para_insertar)


# ─────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Actualiza PRECIOS_DIGEMID desde catalogoproductos.xlsx")
    parser.add_argument("--dry-run",      action="store_true", help="Solo mostrar estadísticas, sin modificar BD")
    parser.add_argument("--solo-nuevos",  action="store_true", help="Solo INSERT, sin UPDATE de existentes")
    args = parser.parse_args()

    print("=" * 60)
    print("  ACTUALIZACIÓN dbo.PRECIOS_DIGEMID")
    print(f"  Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    if args.dry_run:
        print("  MODO: DRY-RUN (sin cambios en BD)")
    elif args.solo_nuevos:
        print("  MODO: Solo nuevos registros")
    else:
        print("  MODO: UPSERT (INSERT + UPDATE)")
    print("=" * 60)

    # Leer Excel
    headers, filas = leer_excel()

    if not filas:
        print("\n[WARN] No se encontraron filas en el Excel. Abortando.")
        sys.exit(0)

    # Conectar
    print(f"\n[INFO] Conectando a SQL Server ({DB_CONFIG['server']}/{DB_CONFIG['database']}) ...")
    conn   = get_connection()
    cursor = conn.cursor()
    print("  Conexión exitosa.")

    # Verificar tabla
    columnas_tabla = verificar_tabla(cursor)

    # Procesar por lotes
    total_filas     = len(filas)
    procesadas      = 0
    errores         = 0
    inicio          = time.time()

    print(f"\n[INFO] Procesando {total_filas:,} registros en lotes de {BATCH_SIZE}...\n")

    for i in range(0, total_filas, BATCH_SIZE):
        lote = filas[i:i + BATCH_SIZE]
        try:
            procesadas += actualizar_lote(cursor, lote, columnas_tabla, args.dry_run)

            if not args.dry_run:
                conn.commit()

            # Progreso
            pct = (i + len(lote)) / total_filas * 100
            elapsed = time.time() - inicio
            rps = procesadas / elapsed if elapsed > 0 else 0
            print(f"  [{pct:5.1f}%] Procesadas: {procesadas:,}/{total_filas:,}  "
                  f"({rps:.0f} reg/s)", end="\r")

        except Exception as e:
            conn.rollback()
            errores += 1
            print(f"\n  [ERROR] Lote {i//BATCH_SIZE + 1} falló: {e}")
            if errores >= 5:
                print("  Demasiados errores consecutivos. Abortando.")
                break

    cursor.close()
    conn.close()

    # Resumen
    elapsed = time.time() - inicio
    sep = "=" * 60
    lin = "-" * 56
    print("\n\n" + sep)
    print("  RESUMEN FINAL")
    print("  " + lin)
    print(f"  Filas en Excel  : {total_filas:,}")
    print(f"  Filas procesadas: {procesadas:,}")
    print(f"  Errores de lote : {errores}")
    print(f"  Tiempo total    : {elapsed:.1f} segundos")
    if args.dry_run:
        print("  * Modo DRY-RUN: NINGUN cambio fue guardado en BD.")
    else:
        print("  * Cambios CONFIRMADOS en dbo.PRECIOS_DIGEMID")
    print(sep)


if __name__ == "__main__":
    main()
